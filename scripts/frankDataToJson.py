#!/usr/bin/env python2.7

import csv
import os
import openpyxl
import re
import json
import sys
import pickle
import pprint

country_code = {
    'Aruba': 'ABW',
    'Andorra': 'AND',
    'Afghanistan': 'AFG',
    'Angola': 'AGO',
    'Anguilla': 'AIA',
    'Albania': 'ALB',
    'Arab World': 'ARB',
    'United Arab Emirates': 'ARE',
    'Argentina': 'ARG',
    'Armenia': 'ARM',
    'American Samoa': 'ASM',
    'Antigua and Barbuda': 'ATG',
    'Australia': 'AUS',
    'Austria': 'AUT',
    'Azerbaijan': 'AZE',
    'Burundi': 'BDI',
    'Belgium': 'BEL',
    'Benin': 'BEN',
    'Burkina Faso': 'BFA',
    'Bangladesh': 'BGD',
    'Bulgaria': 'BGR',
    'Bahrain': 'BHR',
    'Bahamas, The': 'BHS',
    'Bahamas': 'BHS',
    'Bosnia and Herzegovina': 'BIH',
    'Belarus': 'BLR',
    'Belize': 'BLZ',
    'Bermuda': 'BMU',
    'Bolivia': 'BOL',
    'Bolivia (Plurinational State of)': 'BOL',
    'Brazil': 'BRA',
    'Barbados': 'BRB',
    'Brunei Darussalam': 'BRN',
    'Bhutan': 'BTN',
    'Botswana': 'BWA',
    'Central African Republic': 'CAF',
    'Canada': 'CAN',
    'Central Europe and the Baltics': 'CEB',
    'Switzerland': 'CHE',
    'Channel Islands': 'CHI',
    'Chile': 'CHL',
    'China': 'CHN',
    'Cote d\'Ivoire': 'CIV',
    'Cameroon': 'CMR',
    'Congo, Rep.': 'COG',
    'Congo': 'COG',
    'Democratic Republic of the Congo': 'COG',
    'Cook Islands': 'COK',
    'Colombia': 'COL',
    'Comoros': 'COM',
    'Cabo Verde': 'CPV',
    'Costa Rica': 'CRI',
    'Caribbean small states': 'CSS',
    'Cuba': 'CUB',
    'Curacao': 'CUW',
    'Cayman Islands': 'CYM',
    'Cyprus': 'CYP',
    'Czech Republic': 'CZE',
    'Germany': 'DEU',
    'Djibouti': 'DJI',
    'Dominica': 'DMA',
    'Denmark': 'DNK',
    'Dominican Republic': 'DOM',
    'Algeria': 'DZA',
    'East Asia & Pacific (developing only)': 'EAP',
    'East Asia & Pacific (all income levels)': 'EAS',
    'Europe & Central Asia (developing only)': 'ECA',
    'Europe & Central Asia (all income levels)': 'ECS',
    'Ecuador': 'ECU',
    'Egypt, Arab Rep.': 'EGY',
    'Egypt': 'EGY',
    'Euro area': 'EMU',
    'Eritrea': 'ERI',
    'Spain': 'ESP',
    'Estonia': 'EST',
    'Ethiopia': 'ETH',
    'European Union': 'EUU',
    'Fragile and conflict affected situations': 'FCS',
    'Finland': 'FIN',
    'Fiji': 'FJI',
    'France': 'FRA',
    'Faeroe Islands': 'FRO',
    'Micronesia, Fed. Sts.': 'FSM',
    'Micronesia (Federated States of)': 'FSM',
    'Gabon': 'GAB',
    'United Kingdom': 'GBR',
    'United Kingdom of Great Britain and Northern Ireland': 'GBR',
    'Georgia': 'GEO',
    'Ghana': 'GHA',
    'Guinea': 'GIN',
    'Gambia, The': 'GMB',
    'Gambia': 'GMB',
    'Guinea-Bissau': 'GNB',
    'Equatorial Guinea': 'GNQ',
    'Greece': 'GRC',
    'Grenada': 'GRD',
    'Greenland': 'GRL',
    'Guatemala': 'GTM',
    'Guam': 'GUM',
    'Guyana': 'GUY',
    'High income': 'HIC',
    'Hong Kong SAR, China': 'HKG',
    'Honduras': 'HND',
    'Heavily indebted poor countries (HIPC)': 'HPC',
    'Croatia': 'HRV',
    'Haiti': 'HTI',
    'Hungary': 'HUN',
    'Indonesia': 'IDN',
    'Isle of Man': 'IMN',
    'India': 'IND',
    'Not classified': 'INX',
    'Ireland': 'IRL',
    'Iran, Islamic Rep.': 'IRN',
    'Iran (Islamic Republic of)': 'IRN',
    'Iraq': 'IRQ',
    'Iceland': 'ISL',
    'Israel': 'ISR',
    'Italy': 'ITA',
    'Jamaica': 'JAM',
    'Jordan': 'JOR',
    'Japan': 'JPN',
    'Kazakhstan': 'KAZ',
    'Kenya': 'KEN',
    'Kyrgyz Republic': 'KGZ',
    'Kyrgyzstan': 'KGZ',
    'Cambodia': 'KHM',
    'Kiribati': 'KIR',
    'St. Kitts and Nevis': 'KNA',
    'Saint Kitts and Nevis': 'KNA',
    'Korea, Rep.': 'KOR',
    'Republic of Korea': 'KOR',
    'Kosovo': 'KSV',
    'Kuwait': 'KWT',
    'Latin America & Caribbean (developing only)': 'LAC',
    'Lao PDR': 'LAO',
    'Lao People\'s Democratic Republic': 'LAO',
    'Lebanon': 'LBN',
    'Liberia': 'LBR',
    'Libya': 'LBY',
    'St. Lucia': 'LCA',
    'Saint Lucia': 'LCA',
    'Latin America & Caribbean (all income levels)': 'LCN',
    'Least developed countries: UN classification': 'LDC',
    'Low income': 'LIC',
    'Liechtenstein': 'LIE',
    'Sri Lanka': 'LKA',
    'Lower middle income': 'LMC',
    'Low & middle income': 'LMY',
    'Lesotho': 'LSO',
    'Lithuania': 'LTU',
    'Luxembourg': 'LUX',
    'Latvia': 'LVA',
    'Macao SAR, China': 'MAC',
    'St. Martin (French part)': 'MAF',
    'Morocco': 'MAR',
    'Monaco': 'MCO',
    'Moldova': 'MDA',
    'Republic of Moldova': 'MDA',
    'Madagascar': 'MDG',
    'Maldives': 'MDV',
    'Middle East & North Africa (all income levels)': 'MEA',
    'Mexico': 'MEX',
    'Marshall Islands': 'MHL',
    'Middle income': 'MIC',
    'Macedonia, FYR': 'MKD',
    'The former Yugoslav republic of Macedonia': 'MKD',
    'Mali': 'MLI',
    'Malta': 'MLT',
    'Myanmar': 'MMR',
    'Middle East & North Africa (developing only)': 'MNA',
    'Montenegro': 'MNE',
    'Mongolia': 'MNG',
    'Northern Mariana Islands': 'MNP',
    'Mozambique': 'MOZ',
    'Mauritania': 'MRT',
    'Mauritius': 'MUS',
    'Malawi': 'MWI',
    'Malaysia': 'MYS',
    'North America': 'NAC',
    'Namibia': 'NAM',
    'New Caledonia': 'NCL',
    'Niger': 'NER',
    'Nigeria': 'NGA',
    'Nicaragua': 'NIC',
    'Niue': 'NIU',
    'Netherlands': 'NLD',
    'High income: nonOECD': 'NOC',
    'Norway': 'NOR',
    'Nepal': 'NPL',
    'Nauru': 'NRU',
    'New Zealand': 'NZL',
    'High income: OECD': 'OEC',
    'OECD members': 'OED',
    'Oman': 'OMN',
    'Other small states': 'OSS',
    'Pakistan': 'PAK',
    'Panama': 'PAN',
    'Peru': 'PER',
    'Philippines': 'PHL',
    'Palau': 'PLW',
    'Papua New Guinea': 'PNG',
    'Poland': 'POL',
    'Puerto Rico': 'PRI',
    'Korea, Dem. Rep.': 'PRK',
    'Democratic People\'s Republic of Korea': 'PRK',
    'Portugal': 'PRT',
    'Paraguay': 'PRY',
    'Pacific island small states': 'PSS',
    'French Polynesia': 'PYF',
    'Qatar': 'QAT',
    'Romania': 'ROU',
    'Russian Federation': 'RUS',
    'Rwanda': 'RWA',
    'South Asia': 'SAS',
    'Saudi Arabia': 'SAU',
    'The former state union Serbia and Montenegro': 'SCG',
    'Serbia and Montenegro': 'SCG',
    'Sudan': 'SDN',
    'Senegal': 'SEN',
    'Singapore': 'SGP',
    'Solomon Islands': 'SLB',
    'Sierra Leone': 'SLE',
    'El Salvador': 'SLV',
    'San Marino': 'SMR',
    'Somalia': 'SOM',
    'Serbia': 'SRB',
    'Sub-Saharan Africa (developing only)': 'SSA',
    'South Sudan': 'SSD',
    'Sub-Saharan Africa (all income levels)': 'SSF',
    'Small states': 'SST',
    'Sao Tome and Principe': 'STP',
    'Suriname': 'SUR',
    'Slovak Republic': 'SVK',
    'Slovakia': 'SVK',
    'Slovenia': 'SVN',
    'Sweden': 'SWE',
    'Swaziland': 'SWZ',
    'Sint Maarten (Dutch part)': 'SXM',
    'Seychelles': 'SYC',
    'Syrian Arab Republic': 'SYR',
    'Turks and Caicos Islands': 'TCA',
    'Chad': 'TCD',
    'Togo': 'TGO',
    'Thailand': 'THA',
    'Tajikistan': 'TJK',
    'Turkmenistan': 'TKM',
    'Timor-Leste': 'TLS',
    'Tonga': 'TON',
    'Trinidad and Tobago': 'TTO',
    'Tunisia': 'TUN',
    'Turkey': 'TUR',
    'Tuvalu': 'TUV',
    'Tanzania': 'TZA',
    'United Republic of Tanzania': 'TZA',
    'Uganda': 'UGA',
    'Ukraine': 'UKR',
    'Upper middle income': 'UMC',
    'Uruguay': 'URY',
    'United States': 'USA',
    'United States of America': 'USA',
    'Uzbekistan': 'UZB',
    'St. Vincent and the Grenadines': 'VCT',
    'Saint Vincent and the Grenadines': 'VCT',
    'Venezuela, RB': 'VEN',
    'Venezuela (Bolivarian Republic of)': 'VEN',
    'Virgin Islands (U.S.)': 'VIR',
    'Vietnam': 'VNM',
    'Viet Nam': 'VNM',
    'Vanuatu': 'VUT',
    'West Bank and Gaza': 'PSE',
    'World': 'WLD',
    'Samoa': 'WSM',
    'Yemen, Rep.': 'YEM',
    'Yemen': 'YEM',
    'South Africa': 'ZAF',
    'Congo, Dem. Rep.': 'COD',
    'Zambia': 'ZMB',
    'Zimbabwe': 'ZWE'
}


#################################################################
#
# OLD FOLDER
#
#################################################################

path = '../attr_data/WDI/old'
# Array of 242 elements
elements = (252-4)*(2015-1960+1)
data = []
i = 0
for filename in os.listdir(path):
    if '.' not in filename or filename.split('.')[1] != 'xlsx':
        continue
    wb = openpyxl.load_workbook(path + '/' + filename)
    sheet = wb.get_sheet_by_name('Data')
    date = sheet['B2'].value
    for row in range(5, 253):
        for rowOfCellObjects in sheet['E' + str(row):'BH' + str(row)]:
            id = sheet['B' + str(row)].value
            name = sheet['A' + str(row)].value
            attr = sheet['C' + str(row)].value.replace(' ', '_').replace('.', '_').lower()
            if attr == '':
                print 'HOLY COW!!!!'
                print sheet['C' + str(row)].value
            for cellObj in rowOfCellObjects:
                country = {}
                country['id'] = id
                country['name'] = name
                country[attr] = cellObj.value
                out = re.match(r'([A-Z]+)([0-9]+)', cellObj.coordinate)
                country['date'] = int(sheet[out.group(1) + '4'].value)
                if len(data) < elements:
                    sys.stdout.write('Appending: ' + id + ' ' +
                                     filename + '\r')
                    # First file, still processing countries and years
                    if cellObj.value:
                        data.append(country)
                else:
                    sys.stdout.write('Next countries: ' + id +
                                     ' ' + filename + '\r')
                    # Rest of files, find position for country and year
                    for old_country in data:
                        if old_country['id'] == id and old_country['date'] == country['date']:
                            if cellObj.value:
                                old_country[attr] = cellObj.value
    sys.stdout.write('\n')
    i += 1

for old_country in data:
    if old_country['id'] == 'AGO' and old_country['date'] == 2010:
        pprint.pprint(old_country)

# output = open(path + '/my_frank.pkl', 'wb')
# pickle.dump(data, output)
# output.close()
#################################################################
#
# BLUE FOLDER
#
#################################################################
# input = open(path + '/frank.pkl', 'r')
# data = pickle.load()
# input.close()

path = '../attr_data/WDI/blue'
for filename in os.listdir(path):
    if '.' not in filename or filename.split('.')[1] != 'xltx':
        continue
    wb = openpyxl.load_workbook(path + '/' + filename)
    sheet = wb.get_sheet_by_name('Data')
    attr = sheet['B1'].value.replace(' ', '_').replace('.', '_').lower()
    my_row = 0
    for row in sheet.rows:
        my_row += 1
        if my_row < 3:
            continue
        id = country_code[row[0].value]
        sys.stdout.write('BLUE: Next countries: ' + id + ' ' + filename + '\r')
        for year in row[1:len(row)]:
            out = re.match(r'([A-Z]+)([0-9]+)', year.coordinate)
            date = int(sheet[out.group(1) + '2'].value)
            # Rest of files, find position for country and year
            for old_country in data:
                if old_country['id'] == id and old_country['date'] == date:
                    if year.value:
                        old_country[attr] = year.value

    sys.stdout.write('\n')

for old_country in data:
    if old_country['id'] == 'AGO' and old_country['date'] == 2010:
        pprint.pprint(old_country)

# output = open(path + '/frank2.pkl', 'wb')
# pickle.dump(data, output)
# output.close()
#################################################################
#
# GREEN FOLDER
#
#################################################################

# data = pickle.load(open(path + '/frank2.pkl', 'r'))

path = '../attr_data/WDI/green'
for filename in os.listdir(path):
    if '.' not in filename or filename.split('.')[1] != 'xltx':
        continue
    wb = openpyxl.load_workbook(path + '/' + filename)
    sheet = wb.get_sheet_by_name('Data')
    attr = sheet['C1'].value.replace(' ', '_').replace('.', '_').lower()
    my_row = 0
    for row in sheet.rows:
        my_row += 1
        if my_row < 2:
            continue
        if row[0].value is not None:
            id = country_code[row[0].value]
            old_row = id
        else:
            id = old_row
        try:
            date = int(row[1].value)
        except:
            continue
        value = row[2].value
        sys.stdout.write('GREEN: Next countries: ' + id + ' ' + filename + '\r')
        for old_country in data:
            if old_country['id'] == id and old_country['date'] == date:
                old_country[attr] = value
    sys.stdout.write('\n')

# output = open(path + '/old_attr.pkl', 'wb')
#
# # Pickle dictionary using protocol 0.
# pickle.dump(data, output)
#
# g = open(path + '/attr.json', 'w')
# g.write(json.dumps(data))
# g.close()

for old_country in data:
    if old_country['id'] == 'AGO' and old_country['date'] == 2010:
        pprint.pprint(old_country)

#################################################################
#
# PINK FOLDER
#
#################################################################

# data = pickle.load(open(path + '/old_attr.pkl', 'r'))
path = '../attr_data/WDI/pink'

for filename in os.listdir(path):
    if '.' not in filename or filename.split('.')[1] != 'xlsx':
        continue
    wb = openpyxl.load_workbook(path + '/' + filename)
    sheet = wb.get_sheet_by_name('Data')
    my_row = 0
    for row in sheet.rows:
        my_row += 1
        if my_row < 2:
            continue
        # Country ID
        if row[0].value is not None:
            id = country_code[row[0].value]
            old_row = id
        else:
            id = old_row
        # Date attr
        try:
            date = int(row[1].value)
        except:
            continue
        for column in row[2:len(row)]:
            out = re.match(r'([A-Z]+)([0-9]+)', column.coordinate)
            attr = sheet[out.group(1) + '1'].value
            value = row[2].value
            sys.stdout.write('PINK: Next countries: ' + id + ' ' + filename + '\r')
            for old_country in data:
                if old_country['id'] == id and old_country['date'] == date:
                    old_country[attr] = column.value
    sys.stdout.write('\n')

output = open(path + '/old_attr.pkl', 'wb')

# Pickle dictionary using protocol 0.
pickle.dump(data, output)

g = open(path + '/attr.json', 'w')
g.write(json.dumps(data))
g.close()
